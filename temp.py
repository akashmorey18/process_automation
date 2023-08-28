import os
import streamlit as st
import pandas as pd
from datetime import datetime
import streamlit as st
from PIL import Image
from openpyxl import load_workbook

import streamlit as st
from PIL import Image
from office365.runtime.auth.authentication_context import AuthenticationContext
from office365.sharepoint.client_context import ClientContext
from office365.sharepoint.files.file_creation_information import FileCreationInformation

# SharePoint authentication using client ID and client secret

site_url = "https://tatadigitalltd.sharepoint.com/sites/AutomationProject"
client_id = '96026163-d873-4cca-abd4-7acf4fa3e1a3'
client_secret = 'DolnVP4E9IIoLNf/86GHbABu+djbFE679RerpV5Z9Hs='


ctx_auth = AuthenticationContext(url=site_url)
ctx_auth.acquire_token_for_app(client_id, client_secret)
ctx = ClientContext(site_url, ctx_auth)

# Upload a file to SharePoint using the office365 library
def upload_to_sharepoint(file_content, file_name, folder_url):
    ctx_auth = AuthenticationContext(url=site_url)
    ctx_auth.acquire_token_for_app(client_id, client_secret)
    ctx = ClientContext(site_url, ctx_auth)
    
    target_folder = ctx.web.get_folder_by_server_relative_url(folder_url)
    file_info = FileCreationInformation()
    file_info.url = file_name
    target_file = target_folder.files.add(file_info)
    
    # Update the content of the target file
    ctx.execute_query()
    with target_file.file.open_binary_stream() as binary_stream:
        binary_stream.write(file_content)
    
    ctx.execute_query()
    
# Set the page title and header
st.set_page_config(page_title="Fashion & Lifestyle Merchandising")
st.title("Fashion & Lifestyle Merchandising")

# Create the sidebar menu
menu_items = {
    "Offer and Content Creation": 1,
    "Creatives Requests": 2,
    "Creatives Upload": 3,
    "Creatives Approval": 4,
    "Daily Ops Tracker Summary": 5
}
selected_page = st.sidebar.radio("Navigation", list(menu_items.keys()))

# Render the selected page based on the sidebar selection
if selected_page == "Offer and Content Creation":
    st.subheader("Offer and Content Creation")
    # Form fields
    col1, col2 = st.columns(2)
    with col1:
        start_date = st.date_input("Start Date")
    with col2:
        end_date = st.date_input("End Date")
    callout = st.text_input("Offer callout or copy")
    link = st.text_input("Link")
    comments = st.text_input("Comments")

    # Save input to Excel on submit button click
    if st.button("Submit"):
        # Create the folder if it doesn't exist
        folder_path = "AutomationProject"
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)

        # Generate primary key
        file_path = os.path.join(folder_path, "offer_content_creation.xlsx")
        if os.path.isfile(file_path):
            existing_df = pd.read_excel(file_path, sheet_name="Sheet1")
            last_primary_key = existing_df["Primary Key"].str.extract(r'(\d+)').astype(int).max().values
            next_primary_key = last_primary_key + 1 if len(last_primary_key) > 0 else 1
        else:
            next_primary_key = 1

        # Prepare the data
        data = {
            "Primary Key": ["LSMR" + str(next_primary_key).zfill(3)],
            "Start Date": [start_date],
            "End Date": [end_date],
            "Callout": [callout],
            "Link": [link],
            "Comments": [comments],
            "Timestamp": [datetime.now()]
        }
        df = pd.DataFrame(data)

        try:
            # Append or create the Excel file
            if os.path.isfile(file_path):
                with pd.ExcelWriter(file_path, mode='a', engine='openpyxl') as writer:
                    writer.book = load_workbook(file_path)
                    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
                    sheet_name = "Sheet1"
                    if sheet_name in writer.sheets:
                        startrow = writer.sheets[sheet_name].max_row
                        df.to_excel(writer, index=False, header=False, sheet_name=sheet_name, startrow=startrow+1)
                    else:
                        df.to_excel(writer, index=False, header=True, sheet_name=sheet_name)
            else:
                df.to_excel(file_path, index=False, sheet_name="Sheet1")

            st.success("Data saved successfully!")
        except Exception as e:
            st.error(f"An error occurred while saving the data: {e}")
            
    # Display a success message if data is successfully saved


if selected_page == "Creatives Upload" :

    def main():
        st.title("Image Upload to SharePoint")
    
    # Display the uploaded image and allow uploading to SharePoint
    uploaded_image = st.file_uploader("Choose an image", type=["jpg", "png", "jpeg"])
    
    if uploaded_image is not None:
        image = Image.open(uploaded_image)
        st.image(image, caption="Uploaded Image", use_column_width=True)

        if st.button("Upload to SharePoint"):
            image_content = uploaded_image.read()
            folder_url = "/sites/AutomationProject/Shared Documents/Fashion Merchandising"
            file_name = uploaded_image.name
            upload_to_sharepoint(image_content, file_name, folder_url)
            st.success("Image uploaded to SharePoint!")
            
    if __name__ == "__main__":
        main()
