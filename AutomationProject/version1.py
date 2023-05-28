import streamlit as st
import pandas as pd
import openpyxl
import uuid

def save_to_excel(data, sheet_name):
    df = pd.DataFrame(data)
    with pd.ExcelWriter("form_data.xlsx", engine="openpyxl", mode="a") as writer:
        writer.book = openpyxl.load_workbook("form_data.xlsx")
        if sheet_name in writer.book.sheetnames:
            # If sheet exists, remove it before saving the new data
            writer.book.remove(writer.book[sheet_name])
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    st.success("Data saved to form_data.xlsx")

def edit_data(sheet_name):
    try:
        df = pd.read_excel("form_data.xlsx", sheet_name=sheet_name)
        st.dataframe(df)

        st.subheader("Edit Data")
        primary_key = st.text_input("Enter Primary Key of the row to edit")

        if st.button("Edit"):
            mask = df["Primary Key"] == primary_key
            if mask.any():
                row_idx = df.index[mask][0]

                # Display current data
                st.write("Current Data:")
                st.write(df.loc[row_idx])

                # Update data
                new_details = st.text_input("Enter new Details", value=df.loc[row_idx, "Details"])
                new_callout = st.text_input("Enter new Offer Callout or Copy", value=df.loc[row_idx, "Offer Callout or Copy"])
                new_link = st.text_input("Enter new Link", value=df.loc[row_idx, "Link"])
                new_comments = st.text_input("Enter new Comments", value=df.loc[row_idx, "Comments"])

                if st.button("Save Changes"):
                    df.loc[row_idx, "Details"] = new_details
                    df.loc[row_idx, "Offer Callout or Copy"] = new_callout
                    df.loc[row_idx, "Link"] = new_link
                    df.loc[row_idx, "Comments"] = new_comments
                    with pd.ExcelWriter("form_data.xlsx", engine="openpyxl", mode="a") as writer:
                        writer.book = openpyxl.load_workbook("form_data.xlsx")
                        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
                        if sheet_name in writer.sheets:
                            del writer.sheets[sheet_name]
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                    st.success("Data updated successfully")
            else:
                st.warning("No row found with the provided Primary Key")
    except FileNotFoundError:
        st.warning("No data found. Please submit the form first.")

def main():
    st.title("Fashion & Lifestyle Merchandising")

    # Sidebar section
    st.sidebar.title("Sections")
    sections = ["Offer and Content Creation", "Form 2"]
    choice = st.sidebar.radio("Select Section", sections)

    if choice == "Offer and Content Creation":
        st.subheader("Offer and Content Creation")

        # Display all rows in the sheet
        try:
            df = pd.read_excel("form_data.xlsx", sheet_name="Offer and Content Creation")
            st.dataframe(df)
        except FileNotFoundError:
            st.warning("No data found. Please submit the form first.")

        # Form inputs
        details = st.text_area("Details")
        callout = st.text_input("Offer Callout or Copy")
        link = st.text_input("Link")
        comments = st.text_area("Comments")

        if st.button("Submit"):
            # Save form data to Excel (Sheet 1 - Offer and Content Creation)
            data = {
                "Primary Key": [str(uuid.uuid4())[:8]],
                "Details": [details],
                "Offer Callout or Copy": [callout],
                "Link": [link],
                "Comments": [comments]
            }
            save_to_excel(data, "Offer and Content Creation")

        st.subheader("Edit Data")
        edit_data("Offer and Content Creation")

    elif choice == "Form 2":
        st.subheader("Form 2")

        # Form inputs
        name = st.text_input("Name")
        age = st.number_input("Age", min_value=0, max_value=120)
        email = st.text_input("Email")

        if st.button("Submit"):
            # Save form data to Excel (Sheet 2 - Form 2)
            data = {
                "Primary Key": [str(uuid.uuid4())[:8]],
                "Name": [name],
                "Age": [age],
                "Email": [email]
            }
            save_to_excel(data, "Form 2")

        st.subheader("Edit Data")
        edit_data("Form 2")

if __name__ == "__main__":
    main()
